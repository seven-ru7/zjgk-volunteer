"""导出器测试。"""
import pathlib

import pandas as pd
import pytest
import openpyxl

from src.exporter import (
    to_dataframe, to_excel, to_csv,
    to_excel_multi_sheet, get_institution_card_data,
    make_sparkline, compute_trend,
)
from src.models import Recommendation, Program


def make_rec(pid, name, inst, tier, prob, delta, history=None):
    p = Program(
        program_id=pid,
        name=name,
        institution=inst,
        city="杭州",
        duration_years=4,
        tuition=6000,
        required_subjects=["物理"],
        plan_quota_2026=50,
        history=history or [
            {"year": 2025, "min_rank": 12500, "min_score": 632},
            {"year": 2024, "min_rank": 11800, "min_score": 638},
            {"year": 2023, "min_rank": 11000, "min_score": 640},
        ],
    )
    return Recommendation(program=p, tier=tier, probability=prob, delta_rank=delta)


class TestToDataFrame:
    def test_columns(self):
        recs = [make_rec("P1", "CS", "A大学", "稳", 0.75, 1000)]
        df = to_dataframe(recs)
        expected = [
            "序号", "层次", "院校", "专业", "城市", "学制", "学费",
            "2026计划数", "2025最低位次", "2025最低分",
            "2024最低位次", "2023最低位次", "趋势", "3年趋势",
            "位次差", "录取概率",
        ]
        assert list(df.columns) == expected

    def test_values(self):
        recs = [make_rec("P1", "CS", "A大学", "稳", 0.75, 1000)]
        df = to_dataframe(recs)
        row = df.iloc[0]
        assert row["序号"] == 1
        assert row["层次"] == "稳"
        assert row["院校"] == "A大学"
        assert row["录取概率"] == "75%"
        assert row["2025最低位次"] == 12500
        assert row["2025最低分"] == 632
        assert row["3年趋势"] != ""  # 有 sparkline

    def test_multiple_rows(self):
        recs = [
            make_rec("P1", "CS", "A", "冲", 0.45, 5000),
            make_rec("P2", "SE", "B", "稳", 0.75, 0),
            make_rec("P3", "AI", "C", "保", 0.95, -10000),
        ]
        df = to_dataframe(recs)
        assert len(df) == 3
        assert df.iloc[0]["序号"] == 1
        assert df.iloc[2]["序号"] == 3

    def test_no_history(self):
        p = Program(
            program_id="X", name="X", institution="X", city="X",
            duration_years=4, tuition=5000, required_subjects=[],
            plan_quota_2026=10, history=[],
        )
        rec = Recommendation(program=p, tier="稳", probability=0.5, delta_rank=0)
        df = to_dataframe([rec])
        assert pd.isna(df.iloc[0]["2025最低位次"])
        assert df.iloc[0]["趋势"] == "数据不足"


class TestSparkline:
    def test_three_values(self):
        result = make_sparkline([10000, 11000, 12000])
        assert len(result) == 3
        # 位次递增（变容易）→ block 高度应该递减
        # 最高位次 10000 = 最难 = 最高 block (█)
        # 最低位次 12000 = 最易 = 最低 block (▁)
        assert result[0] == "█"  # 最高 block
        assert result[2] == "▁"  # 最低 block

    def test_stable(self):
        result = make_sparkline([10000, 10000, 10000])
        # 完全相同 → 中间 block
        assert "▅" in result

    def test_missing_values(self):
        result = make_sparkline([None, 10000, 12000])
        assert result[0] == "░"
        assert len(result) == 3

    def test_single_value(self):
        result = make_sparkline([None, None, 10000])
        assert "░" in result


class TestTrend:
    def test_stable(self):
        t = compute_trend([10000, 10000, 10000])
        assert t["direction"] == "→"
        assert t["stable"] is True

    def test_getting_easier(self):
        t = compute_trend([10000, 11000, 12000])
        assert t["direction"] == "↓"
        assert t["delta"] == 2000
        assert t["delta_pct"] == 20.0

    def test_getting_harder(self):
        t = compute_trend([12000, 11000, 10000])
        assert t["direction"] == "↑"
        assert t["delta"] == -2000

    def test_missing_data(self):
        t = compute_trend([None, None, None])
        assert t["direction"] == "—"
        t = compute_trend([10000, None, None])
        assert t["direction"] == "—"

    def test_small_change_stable(self):
        t = compute_trend([10000, 10300, 10400])
        # +4% < 5% → 平稳
        assert t["direction"] == "→"


class TestToExcel:
    def test_creates_file(self, tmp_path):
        recs = [make_rec("P1", "CS", "A大学", "稳", 0.75, 1000)]
        out = to_excel(recs, tmp_path / "test.xlsx")
        assert out.exists()
        assert out.stat().st_size > 0
        df = pd.read_excel(out, engine="openpyxl")
        assert len(df) == 1
        assert "3年趋势" in df.columns


class TestToCsv:
    def test_creates_file(self, tmp_path):
        recs = [make_rec("P1", "CS", "A大学", "稳", 0.75, 1000)]
        out = to_csv(recs, tmp_path / "test.csv")
        assert out.exists()
        content = out.read_text(encoding="utf-8-sig")
        assert "A大学" in content
        assert "75%" in content
        assert "3年趋势" in content


class TestToExcelMultiSheet:
    def test_creates_4_sheets(self, tmp_path):
        recs = [
            make_rec("R1", "CS", "A", "冲", 0.45, 5000),
            make_rec("R2", "SE", "B", "冲", 0.45, 4000),
            make_rec("S1", "Math", "C", "稳", 0.75, 0),
            make_rec("S2", "Phys", "D", "稳", 0.75, 1000),
            make_rec("S3", "Chem", "E", "稳", 0.75, -1000),
            make_rec("P1", "Ag", "F", "保", 0.95, -10000),
        ]
        candidate = {
            "score": 620, "rank": 32114,
            "subjects": ["物理", "化学", "生物"],
            "cities": [], "keywords": [],
        }
        out = to_excel_multi_sheet(recs, tmp_path / "multi.xlsx", candidate=candidate)
        assert out.exists()
        assert out.stat().st_size > 0

        wb = openpyxl.load_workbook(out)
        # 应有 4 个 sheet：摘要 + 冲 + 稳 + 保
        assert len(wb.sheetnames) == 4
        assert any("摘要" in s for s in wb.sheetnames)
        assert any("冲档" in s for s in wb.sheetnames)
        assert any("稳档" in s for s in wb.sheetnames)
        assert any("保档" in s for s in wb.sheetnames)

    def test_summary_contains_candidate_info(self, tmp_path):
        recs = [make_rec("S1", "Math", "C", "稳", 0.75, 0)]
        candidate = {
            "score": 620, "rank": 32114,
            "subjects": ["物理", "化学", "生物"],
            "cities": ["杭州"], "keywords": ["计算机"],
        }
        out = to_excel_multi_sheet(recs, tmp_path / "test.xlsx", candidate=candidate)
        wb = openpyxl.load_workbook(out)
        summary = wb["📊 摘要"]
        # 检查考生分数、位次、选科、偏好都写入
        text = "\n".join(str(c.value) for row in summary.iter_rows() for c in row if c.value)
        assert "620" in text
        assert "32,114" in text
        assert "物理" in text
        assert "杭州" in text
        assert "计算机" in text


class TestInstitutionCard:
    def test_stability_stable(self):
        rec = make_rec("P", "CS", "A大学", "稳", 0.75, 0, history=[
            {"year": 2025, "min_rank": 12000, "min_score": 632},
            {"year": 2024, "min_rank": 11800, "min_score": 638},
            {"year": 2023, "min_rank": 11500, "min_score": 640},
        ])
        card = get_institution_card_data(rec.program)
        assert card["stability"] == "稳定"
        assert card["cv"] < 0.10

    def test_stability_volatile(self):
        rec = make_rec("P", "CS", "A大学", "稳", 0.75, 0, history=[
            {"year": 2025, "min_rank": 8000, "min_score": 670},
            {"year": 2024, "min_rank": 15000, "min_score": 620},
            {"year": 2023, "min_rank": 50000, "min_score": 580},
        ])
        card = get_institution_card_data(rec.program)
        assert "波动" in card["stability"]

    def test_3y_min_score(self):
        rec = make_rec("P", "CS", "A大学", "稳", 0.75, 0)
        card = get_institution_card_data(rec.program)
        assert len(card["min_score_3y"]) == 3
        # 应该是 (分数, 位次) 元组列表
        for score, rank in card["min_score_3y"]:
            assert isinstance(score, int)
            assert isinstance(rank, int)
