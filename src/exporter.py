"""志愿表导出增强版。

新增功能：
- 多 Sheet Excel 导出（每个 tier 一个 sheet + 摘要 sheet）
- CSV/Excel 包含 sparkline 字符（兼容 Excel）
"""
from __future__ import annotations

import math
import pathlib
from typing import Sequence

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from .models import Program, Recommendation


COLUMNS = [
    "序号", "层次", "院校", "专业", "城市", "学制", "学费",
    "2026计划数", "2025最低位次", "2025最低分", "2024最低位次", "2023最低位次",
    "趋势", "3年趋势",
    "位次差", "录取概率",
]


def _get_year_rank(program: Program, year: int):
    for h in program.history:
        if h.get("year") == year:
            return h.get("min_rank")
    return None


def _get_year_score(program: Program, year: int):
    for h in program.history:
        if h.get("year") == year:
            return h.get("min_score")
    return None


_SPARKLINE_CHARS = "▁▂▃▄▅▆▇█"


def make_sparkline(ranks: list, reverse: bool = True) -> str:
    valid = [(i, r) for i, r in enumerate(ranks) if r is not None and r > 0]
    if not valid:
        return "░░░"
    if len(valid) == 1:
        return "▅░░"
    vals = [r for _, r in valid]
    vmin, vmax = min(vals), max(vals)
    if vmax == vmin:
        return "".join(_SPARKLINE_CHARS[4] if r is not None else "░" for r in ranks)
    blocks = []
    for r in ranks:
        if r is None:
            blocks.append("░")
        else:
            ratio = (r - vmin) / (vmax - vmin)
            if reverse:
                ratio = 1 - ratio
            idx = int(ratio * (len(_SPARKLINE_CHARS) - 1))
            idx = max(0, min(len(_SPARKLINE_CHARS) - 1, idx))
            blocks.append(_SPARKLINE_CHARS[idx])
    return "".join(blocks)


def compute_trend(ranks: list) -> dict:
    if not ranks or ranks[0] is None or ranks[-1] is None:
        return {"direction": "—", "delta": 0, "delta_pct": 0.0, "stable": False}
    r_old, r_new = ranks[0], ranks[-1]
    delta = r_new - r_old
    delta_pct = (delta / r_old * 100) if r_old > 0 else 0
    if abs(delta_pct) < 5:
        direction, stable = "→", True
    elif delta > 0:
        direction, stable = "↓", False
    else:
        direction, stable = "↑", False
    return {"direction": direction, "delta": delta, "delta_pct": round(delta_pct, 1), "stable": stable}


def to_dataframe(recs: Sequence[Recommendation]) -> pd.DataFrame:
    rows = []
    for i, r in enumerate(recs, start=1):
        p = r.program
        h25 = _get_year_rank(p, 2025)
        h24 = _get_year_rank(p, 2024)
        h23 = _get_year_rank(p, 2023)
        h25_score = _get_year_score(p, 2025)
        ranks = [h23, h24, h25]
        trend = compute_trend(ranks)
        sparkline = make_sparkline(ranks, reverse=True)
        if trend["direction"] == "—":
            trend_text = "数据不足"
        elif trend["direction"] == "→":
            trend_text = f"→ 平稳（{trend['delta_pct']:+.1f}%）"
        elif trend["direction"] == "↓":
            trend_text = f"↓ 变易（{trend['delta_pct']:+.1f}%）"
        else:
            trend_text = f"↑ 变难（{trend['delta_pct']:+.1f}%）"
        rows.append({
            "序号": i,
            "层次": r.tier,
            "院校": p.institution,
            "专业": p.name,
            "城市": p.city,
            "学制": p.duration_years,
            "学费": p.tuition,
            "2026计划数": p.plan_quota_2026,
            "2025最低位次": h25,
            "2025最低分": h25_score,
            "2024最低位次": h24,
            "2023最低位次": h23,
            "趋势": trend_text,
            "3年趋势": sparkline,
            "位次差": r.delta_rank,
            "录取概率": f"{r.probability * 100:.0f}%",
        })
    return pd.DataFrame(rows, columns=COLUMNS)


def to_excel(recs: Sequence[Recommendation], path: str | pathlib.Path) -> pathlib.Path:
    """导出为单 sheet Excel。"""
    df = to_dataframe(recs)
    path = pathlib.Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, engine="openpyxl")
    return path


def to_excel_multi_sheet(
    recs: Sequence[Recommendation],
    path: str | pathlib.Path,
    candidate: dict | None = None,
) -> pathlib.Path:
    """导出为多 Sheet Excel：
    - 「📊 摘要」：考生信息 + 各 tier 统计
    - 「🔴 冲档 (N)」：冲档志愿
    - 「🟡 稳档 (N)」：稳档志愿
    - 「🟢 保档 (N)」：保档志愿

    Args:
        recs: 推荐列表
        path: 输出文件路径
        candidate: 考生信息（用于摘要 sheet）
    """
    df = to_dataframe(recs)
    path = pathlib.Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # 颜色样式
    tier_colors = {
        "冲": "FFCCCC",  # 红
        "稳": "FFF3CD",  # 黄
        "保": "D4EDDA",  # 绿
    }
    tier_fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in tier_colors.items()}

    # tier → emoji
    tier_emoji = {"冲": "🔴", "稳": "🟡", "保": "🟢"}

    # 按 tier 分组
    tier_dfs = {}
    for tier in ["冲", "稳", "保"]:
        tier_dfs[tier] = df[df["层次"] == tier].copy()

    # 用 openpyxl 直接写多 sheet（pd.ExcelWriter 也能，但样式控制更弱）
    from openpyxl import Workbook
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # 1. 摘要 sheet
    ws_summary = wb.create_sheet("📊 摘要")
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40

    summary_data = []
    if candidate:
        summary_data.append(("考生分数", f"{candidate.get('score', '?')}"))
        summary_data.append(("对应位次", f"{candidate.get('rank', '?'):,}" if isinstance(candidate.get('rank'), int) else candidate.get('rank', '?')))
        summary_data.append(("选考科目", "、".join(candidate.get("subjects", []))))
        if candidate.get("cities"):
            summary_data.append(("偏好城市", "、".join(candidate["cities"])))
        if candidate.get("keywords"):
            summary_data.append(("专业关键词", "、".join(candidate["keywords"])))
    summary_data.extend([
        ("", ""),
        ("总志愿数", len(df)),
        ("冲档志愿", len(tier_dfs["冲"])),
        ("稳档志愿", len(tier_dfs["稳"])),
        ("保档志愿", len(tier_dfs["保"])),
        ("平均概率", f"{df['录取概率'].apply(lambda x: int(x.rstrip('%'))).mean():.0f}%" if len(df) > 0 else "—"),
        ("", ""),
        ("生成时间", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("数据来源", "浙江省教育考试院（zjzs.net）2023-2025 真实数据"),
    ])
    bold = Font(bold=True)
    for i, (k, v) in enumerate(summary_data, start=1):
        ws_summary.cell(row=i, column=1, value=k).font = bold if k else Font()
        ws_summary.cell(row=i, column=2, value=v)

    # 2. 每个 tier 一个 sheet
    for tier in ["冲", "稳", "保"]:
        tdf = tier_dfs[tier]
        sheet_name = f"{tier_emoji[tier]} {tier}档 ({len(tdf)})"
        # Excel sheet 名最长 31 字符
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(sheet_name)

        # 写表头
        for col_idx, col_name in enumerate(tdf.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = bold
            cell.fill = tier_fills[tier]
            cell.alignment = Alignment(horizontal="center")

        # 写数据
        for row_idx, row in enumerate(tdf.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else "")

        # 列宽自适应
        for col_idx, col_name in enumerate(tdf.columns, start=1):
            col_letter = get_column_letter(col_idx)
            # 中文字符宽度估算
            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in tdf.iloc[:, col_idx - 1] if pd.notna(v)] + [8]
            )
            ws.column_dimensions[col_letter].width = min(max_len * 1.5, 30)

        # 冻结首行
        ws.freeze_panes = "A2"

    # 保存
    wb.save(path)
    return path


def to_csv(recs: Sequence[Recommendation], path: str | pathlib.Path) -> pathlib.Path:
    df = to_dataframe(recs)
    path = pathlib.Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def get_institution_card_data(program: Program) -> dict:
    """获取学校详情卡片数据。"""
    h25 = _get_year_rank(program, 2025)
    h24 = _get_year_rank(program, 2024)
    h23 = _get_year_rank(program, 2023)

    # 3 年趋势
    ranks = [h23, h24, h25]
    trend = compute_trend(ranks)

    # 录取稳定性（变异系数）
    valid = [r for r in ranks if r is not None and r > 0]
    if len(valid) >= 2:
        avg = sum(valid) / len(valid)
        variance = sum((r - avg) ** 2 for r in valid) / len(valid)
        cv = math.sqrt(variance) / avg if avg > 0 else 0
        stability = "稳定" if cv < 0.10 else ("小幅波动" if cv < 0.25 else "大幅波动")
    else:
        stability = "数据不足"
        cv = 0

    return {
        "trend": trend,
        "stability": stability,
        "cv": cv,
        "sparkline": make_sparkline(ranks),
        "min_score_3y": [
            (_get_year_score(program, y), _get_year_rank(program, y))
            for y in [2023, 2024, 2025] if _get_year_rank(program, y)
        ],
    }